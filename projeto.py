# ============================================================
# RELATÓRIO AUTOMÁTICO DE VENDAS
#
# Autor: Luciano
#
# Objetivo:
# Ler automaticamente todas as planilhas de vendas da pasta,
# consolidar os dados, tratar inconsistências, gerar resumos
# e criar um relatório profissional em Excel.
# ============================================================


# ===============================
# IMPORTAÇÃO DAS BIBLIOTECAS
# ===============================

import pandas as pd
import glob

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter



# ============================================================
# 1 - LOCALIZAR AUTOMATICAMENTE TODAS AS PLANILHAS
# ============================================================

# Procura todos os arquivos iniciados por "vendas_"

arquivos = glob.glob("vendas_*.xlsx")



# Lista onde serão armazenados todos os DataFrames
lista_dfs = []



# ============================================================
# 2 - LER TODAS AS PLANILHAS
# ============================================================

for arquivo in arquivos:

    # Lê o arquivo

    df = pd.read_excel(arquivo)

    # Cria uma coluna informando de qual arquivo o registro veio

    df["Arquivo de Origem"] = arquivo

    # Guarda o DataFrame

    lista_dfs.append(df)



# ============================================================
# 3 - JUNTAR TODOS OS DATAFRAMES
# ============================================================

dados = pd.concat(lista_dfs, ignore_index=True)



# ============================================================
# 4 - TRATAR DADOS AUSENTES
# ============================================================

# As linhas que possuem alguma informação ausente serão
# enviadas para uma aba chamada Inconsistências.

inconsistencias = dados[dados.isnull().any(axis=1)]


# Remove registros incompletos da base principal

dados = dados.dropna()



# ============================================================
# 5 - REMOVER DUPLICIDADES
# ============================================================

dados = dados.drop_duplicates()



# ============================================================
# 6 - GARANTIR O TIPO DOS DADOS
# ============================================================

dados["Quantidade"] = pd.to_numeric(dados["Quantidade"])

dados["Valor Unitário"] = pd.to_numeric(dados["Valor Unitário"])



# ============================================================
# 7 - CRIAR A COLUNA TOTAL
# ============================================================

dados["Total"] = dados["Quantidade"] * dados["Valor Unitário"]



# ============================================================
# 8 - GERAR RESUMO POR VENDEDOR
# ============================================================

resumo_vendedores = (

    dados

    .groupby("Vendedor")

    .agg(
        Quantidade_de_Vendas=("Quantidade", "sum"),
        Total_Vendido=("Total", "sum")
    )

    .reset_index()

)



# ============================================================
# 9 - GERAR RESUMO POR PRODUTO
# ============================================================

resumo_produtos = (

    dados

    .groupby("Produto")

    .agg(
        Quantidade_Vendida=("Quantidade", "sum"),
        Faturamento=("Total", "sum")
    )

    .reset_index()

)



# ============================================================
# 10 - EXPORTAR PARA O EXCEL
# ============================================================

with pd.ExcelWriter("relatorio_final.xlsx",
                    engine="openpyxl") as escritor:

    dados.to_excel(
        escritor,
        sheet_name="Dados Consolidados",
        index=False
    )

    resumo_vendedores.to_excel(
        escritor,
        sheet_name="Resumo por Vendedor",
        index=False
    )

    resumo_produtos.to_excel(
        escritor,
        sheet_name="Resumo por Produto",
        index=False
    )

    inconsistencias.to_excel(
        escritor,
        sheet_name="Inconsistências",
        index=False
    )



# ============================================================
# 11 - ABRIR O ARQUIVO PARA FORMATAÇÃO
# ============================================================

planilha = load_workbook("relatorio_final.xlsx")



# ============================================================
# 12 - FORMATAÇÃO DAS ABAS
# ============================================================

for aba in planilha.worksheets:

    # Cabeçalho azul

    for celula in aba[1]:

        celula.fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        celula.font = Font(
            color="FFFFFF",
            bold=True
        )

        celula.alignment = Alignment(horizontal="center")

    # Ajuste automático das colunas

    for coluna in aba.columns:

        maior = 0

        letra = get_column_letter(coluna[0].column)

        for celula in coluna:

            if celula.value:

                maior = max(maior, len(str(celula.value)))

        aba.column_dimensions[letra].width = maior + 3

# ============================================================
# 13 - FORMATAÇÃO DA DATA
# ============================================================

aba = planilha["Dados Consolidados"]

# Procura a coluna chamada Data
coluna_data = None

for celula in aba[1]:
    if celula.value == "Data":
        coluna_data = celula.column
        break

# Formata todas as datas
if coluna_data:
    for linha in range(2, aba.max_row + 1):
        aba.cell(row=linha,
                 column=coluna_data).number_format = "dd/mm/yyyy"
        
aba = planilha["Inconsistências"]

# Procura a coluna chamada Data
coluna_data = None

for celula in aba[1]:
    if celula.value == "Data":
        coluna_data = celula.column
        break

# Formata todas as datas
if coluna_data:
    for linha in range(2, aba.max_row + 1):
        aba.cell(row=linha,
                 column=coluna_data).number_format = "dd/mm/yyyy"



# ============================================================
# 14 - FORMATAÇÃO DE MOEDA
# ============================================================

# ============================================================
# FORMATAÇÃO DE MOEDA
# ============================================================

campos_monetarios = {
    "Dados Consolidados": ["Valor Unitário", "Total"],
    "Resumo por Vendedor": ["Total_Vendido"],
    "Resumo por Produto": ["Faturamento"]
}

for nome_aba, colunas in campos_monetarios.items():

    aba = planilha[nome_aba]

    # Localiza quais colunas devem receber formatação
    mapa_colunas = {}

    for celula in aba[1]:
        if celula.value in colunas:
            mapa_colunas[celula.value] = celula.column

    # Aplica formato monetário somente nas colunas encontradas
    for coluna in mapa_colunas.values():

        for linha in range(2, aba.max_row + 1):

            celula = aba.cell(row=linha, column=coluna)

            if isinstance(celula.value, (int, float)):
                celula.number_format = 'R$ #,##0.00'



# ============================================================
# 15 - CRIAR O GRÁFICO
# ============================================================

aba_resumo = planilha["Resumo por Vendedor"]

grafico = BarChart()

grafico.title = "Total vendido por vendedor"

grafico.y_axis.title = "Valor vendido (R$)"

grafico.x_axis.title = "Vendedor"

grafico.y_axis.majorGridlines = None

valores = Reference(

    aba_resumo,

    min_col=3,

    min_row=1,

    max_row=aba_resumo.max_row

)



nomes = Reference(

    aba_resumo,

    min_col=1,

    min_row=2,

    max_row=aba_resumo.max_row

)



grafico.add_data(

    valores,

    titles_from_data=True

)

grafico.set_categories(nomes)

grafico.height = 12

grafico.width = 20

aba_resumo.add_chart(grafico, "E2")



# ============================================================
# 16 - SALVAR O ARQUIVO
# ============================================================

planilha.save("relatorio_final.xlsx")



# ============================================================
# 17 - RESPONDER ÀS PERGUNTAS
# ============================================================

ranking = resumo_vendedores.sort_values(
    by="Total_Vendido",
    ascending=False
).reset_index(drop=True)


primeiro = ranking.iloc[0]

segundo = ranking.iloc[1]


diferenca = primeiro["Total_Vendido"] - segundo["Total_Vendido"]


ticket_medio = (
    primeiro["Total_Vendido"] /
    primeiro["Quantidade_de_Vendas"]
)



print("=" * 50)

print("RELATÓRIO CONCLUÍDO COM SUCESSO!")

print("=" * 50)

print()

print(f"Maior vendedor: {primeiro['Vendedor']}")

print(f"Total vendido: R$ {primeiro['Total_Vendido']:.2f}")

print()

print(f"Diferença para o segundo colocado: R$ {diferenca:.2f}")

print()

print(
    f"Quantidade vendida: "
    f"{primeiro['Quantidade_de_Vendas']}"
)

print(
    f"Ticket médio: "
    f"R$ {ticket_medio:.2f}"
)

print()

print(
    "Para identificar se o maior resultado veio de "
    "mais vendas ou de vendas mais caras, compare o "
    "ticket médio do vencedor com o dos demais vendedores."
)

print()

print("Arquivo gerado: relatorio_final.xlsx")